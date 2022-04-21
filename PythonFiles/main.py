# This is a sample Python script.
# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.
import os
import openpyxl
from openpyxl import Workbook, load_workbook

path = r"C:/Users/Colin Janowicz/Documents/GitHub/ParseLoadTimesPython/Bitbar Using Filter Logs/3fd75eb26459706c-logs/"
os.chdir(path)

def create_spreadsheet():
    wb = Workbook()
    # grab the active worksheet
    ws = wb.active
    # Data can be assigned directly to cells
    ws['A1'] = "Process"
    # Rows can also be appended
    ws.append([1, 2, 3])
    # Python types will automatically be converted
    ws['A2'] = datetime.datetime.now()
    # Save the file
    wb.save(f"C:/Users/Colin Janowicz/Documents/GitHub/ParseLoadTimesPython/write_sample.xlsx")


def read_files(file_path):
    with open(file_path, 'r') as file:
        # print(file.read())
        print(file.name)



def iterate_over_files():
    print("File name:")
    for file in os.listdir():
        if file.endswith('.log'):
            file_path = f"{path}/{file}"
            read_files(file_path)


def print_hi(name):
    # Use a breakpoint in the code line below to debug your script.
    print(f'Hi, {name}')  # Press Ctrl+F8 to toggle the breakpoint.


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    #print_hi('PyCharm')
    iterate_over_files()

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
